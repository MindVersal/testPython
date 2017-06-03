from grab import Grab
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side
import datetime

print('Grabbing text from temp_report.html\n')

wb = Workbook()
ws = wb.active
rows = []


def grad_noc():
    """
    Login to NOC
    Password asking from console in manual mode
    Response report and save request to file

    :return:
    """
    url_name_login = 'http://10.166.0.27/db/index.php'
    url_name_report = 'http://10.166.0.27/db/index.php?a=repAdmin&c=sendForm&r=equipInventory&id_pop=3533&id_equip_type=&id_port_type='
    g = Grab(timeout=100, connect_timeout=30)
    g.go(url_name_login)
    g.doc.set_input('login', 'pia@ex.tchercom.ru')
    print('Insert password:')
    password = input()
    g.doc.set_input('password', password)
    g.doc.submit()
    print('Logging to Noc is finished')
    print('Please wait about a minute for request NOC ...')
    resp = g.go(url_name_report)
    resp.charset = 'cp1251'
    print('Saving respond.')
    open('./temp_report.html', 'w').write(resp.unicode_body())
    print('Grabbing report from NOC have finished')


def grab_data():
    """
    Procedure for grabbing data from html report and preparing data for excel.
    RegExp:
        Итого --- find word
        (.|\n|\r|\s)*? --- skipping any chars and words
        ADSL2\+ --- find next word
        \s+ --- skipping space char
        (.+\n){1} --- skipping one row
        .+\s+ --- skipping any and space chars
        (\d+) --- take digital chars
        \s+ --- skipping space chars

        Итого --- find word
        (.|\n|\r|\s)*? --- skipping any chars and words
        GPON --- find next word
        \s+ --- skipping space char
        (.+\n){2} --- skipping two rows
        .+ --- skipping any chars
        \((\d+)\) --- take digital chars in parenthesis
        \s+ --- skipping space chars

    :return:
    """
    g = Grab()
    url_name = r'file://temp_report.html'
    g.go(url_name, charset='cp1251')
    # temp_row = ['A#', 'B#', 'C#', 'D#', 'E#']
    temp_row_blank = ['', '', '', '', '']
    temp_row = ['Тип портов', 'Монтировано портов', 'Активных портов', 'Задействовано портов', 'Свободно портов']
    rows.append(temp_row)
    temp_row =[]
    temp_row.append('ADSL2+:')
    temp_row.append(g.doc.rex_search(r'Итого(.|\n|\r|\s)*?ADSL2\+\s+(.+\n){1}.+\s+(\d+)\s+').group(3))
    temp_row.append(g.doc.rex_search(r'Итого(.|\n|\r|\s)*?ADSL2\+\s+(.+\n){3}.+\s+(\d+)\s+').group(3))
    temp_row.append(g.doc.rex_search(r'Итого(.|\n|\r|\s)*?ADSL2\+\s+(.+\n){5}.+\s+(\d+)\s+').group(3))
    temp_row.append(g.doc.rex_search(r'Итого(.|\n|\r|\s)*?ADSL2\+\s+(.+\n){7}.+\s+(\d+)\s+').group(3))
    rows.append(temp_row)
    temp_row = []
    temp_row.append('GPON:')
    temp_row.append(g.doc.rex_search(r'Итого(.|\n|\r|\s)*?GPON\s+(.+\n){2}.+\((\d+)\)\s+').group(3))
    temp_row.append(g.doc.rex_search(r'Итого(.|\n|\r|\s)*?GPON\s+(.+\n){4}.+\((\d+)\)\s+').group(3))
    temp_row.append(g.doc.rex_search(r'Итого(.|\n|\r|\s)*?GPON\s+(.+\n){6}.+\((\d+)\)\s+').group(3))
    temp_row.append(g.doc.rex_search(r'Итого(.|\n|\r|\s)*?GPON\s+(.+\n){7}.+\s+(\d+)\s+').group(3))
    rows.append(temp_row)
    rows.append(temp_row_blank)
    rows.append(temp_row_blank)
    rows.append(temp_row_blank)
    temp_row = ['Абонентов на 1T3 коструктивах', '', '', 'Монтировано на 1T3 коструктивах', '']
    rows.append(temp_row)
    temp_row = ['АТС', 'Кол-во абонентов', '', 'АТС', 'Кол-во портов']
    rows.append(temp_row)
    name_ats = ['31', '26']
    for ats in name_ats:
        temp_row = []
        temp_row.append('8202' + ats)
        temp_row.append(int(g.doc.rex_search(r'8202' + str(ats) + r'(.|\n|\r|\s)*?1T3(.|\n|\r|\s)*?GPON(.|\n|\r|\s)*?' +
                                             r'summary(.*\n*){6}\s+\d+\((\d+)\)\s+').group(5)))
        temp_row.append('')
        temp_row.append('8202' + ats)
        temp_row.append(int(g.doc.rex_search(r'8202' + str(ats) + r'(.|\n|\r|\s)*?1T3(.|\n|\r|\s)*?GPON(.|\n|\r|\s)*?' +
                                             r'summary(.*\n*){2}\s+\d+\((\d+)\)\s+').group(5)))
        rows.append(temp_row)
    temp_row = ['Итого: ', '=SUM(B9:B10)', '', 'Итого: ', '=SUM(E9:E10)']
    rows.append(temp_row)
    rows.append(temp_row_blank)
    rows.append(temp_row_blank)
    rows.append(temp_row_blank)
    temp_row = ['Абонентов на АТС 26', '', '', '', '']
    rows.append(temp_row)
    temp_row = ['OLT', 'Монтировано портов', 'Активных портов', 'Задействовано портов', 'Свободно портов']
    rows.append(temp_row)
    name_olt = ['MA5680T-1', 'MA5680T-2', 'MA5680T-3', 'MA5680T-4', '1T3-1']
    for olt in name_olt:
        temp_row = []
        temp_row.append(olt)
        temp_row.append(int(g.doc.rex_search(r'820226(.|\n|\r|\s)*?' + olt + '(.|\n|\r|\s)*?GPON(.|\n|\r|\s)*?' +
                                             r'summary(.*\n*){2}\s+\d+\((\d+)\)\s+').group(5)))
        temp_row.append(int(g.doc.rex_search(r'820226(.|\n|\r|\s)*?' + olt + '(.|\n|\r|\s)*?GPON(.|\n|\r|\s)*?' +
                                             r'summary(.*\n*){4}\s+\d+\((\d+)\)\s+').group(5)))
        temp_row.append(int(g.doc.rex_search(r'820226(.|\n|\r|\s)*?' + olt + '(.|\n|\r|\s)*?GPON(.|\n|\r|\s)*?' +
                                             r'summary(.*\n*){6}\s+\d+\((\d+)\)\s+').group(5)))
        temp_row.append(int(g.doc.rex_search(r'820226(.|\n|\r|\s)*?' + olt + '(.|\n|\r|\s)*?GPON(.|\n|\r|\s)*?' +
                                             r'summary(.*\n*){8}\s+(\d+)\s+').group(5)))
        rows.append(temp_row)
    temp_row = ['Итого: ', '=SUM(B17:B21)', '=SUM(C17:C21)', '=SUM(D17:D21)', '=SUM(E17:E21)']
    rows.append(temp_row)
    print('Grabbing data have finished.')


def convert_data():
    """
    Convert data to excel workbook

    :return:
    """
    for row in rows:
        ws.append(row)
    print('Converting data have finished.')


def design_data():
    """
    Designing data in excel workbook

    :return:
    """
    ws.merge_cells('A7:B7')
    ws.merge_cells('D7:E7')
    ws.merge_cells('A15:E15')
    style_alignment = Alignment(horizontal='center')
    style_text_bold = Font(bold=True)
    style_border = Border(
        left=Side(border_style='thin', color='FF000000'),
        right=Side(border_style='thin', color='FF000000'),
        top=Side(border_style='thin', color='FF000000'),
        bottom=Side(border_style='thin', color='FF000000')
    )
    style_no_border = Border(
        left=Side(border_style='thin', color='FF000000'),
        right=Side(border_style='thin', color='FF000000'),
        top=Side(border_style=None, color='FF000000'),
        bottom=Side(border_style=None, color='FF000000')
    )
    cells_bold = ['A1',  'B1',  'C1',  'D1',  'E1',
                  'A7',  'B7',         'D7',  'E7',
                  'A8',  'B8',         'D8',  'E8',
                  'A11', 'B11',        'D11', 'E11',
                  'A15', 'B15', 'C15', 'D15', 'E15',
                  'A16', 'B16', 'C16', 'D16', 'E16',
                  'A22', 'B22', 'C22', 'D22', 'E22']
    for cell in cells_bold:
        ws[cell].font = style_text_bold
        ws[cell].alignment = style_alignment
        ws[cell].border = style_border
    cols = ['A', 'B', 'C', 'D', 'E']
    for col in cols:
        for row in range(2, 4):
            ws[col+str(row)].alignment = style_alignment
            ws[col + str(row)].border = style_border
        for row in range(9, 11):
            ws[col+str(row)].alignment = style_alignment
            ws[col + str(row)].border = style_border
        for row in range(17, 22):
            ws[col+str(row)].alignment = style_alignment
            ws[col + str(row)].border = style_border
    ws['C9'].border = style_no_border
    ws['C10'].border = style_no_border
    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 21
    ws.column_dimensions['E'].width = 17
    print('Designing data have finished.')


def save_data():
    """
    Save excel workbook into file and adding data

    :return:
    """
    month_year = datetime.datetime.now().strftime('%d%m%y')
    wb.save('report_' + month_year + '.xlsx')
    print('Saving data have finished.')

grad_noc()
grab_data()
convert_data()
design_data()
save_data()

print('\nTHE END.')
